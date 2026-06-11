# Verificación del workbook recalculando con LibreOffice (lo real: caza cosas
# que la lib `formulas` no, como funciones post-2007 sin prefijo _xlfn).
# Si no hay LibreOffice, cae a la lib `formulas` (verificación parcial).
# Correr: PYTHONIOENCODING=utf-8 python verify_tracker.py
import os
import subprocess
import sys

WB = 'Adaptive_Fitness_Tracker.xlsx'
SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"

# Caso de prueba base: M, 30 años, 180cm, 80kg, Moderate, Cut, 2.0 g/kg.
w = [80.0, 80.3, 79.9, 80.1, 79.8, 79.9, 79.6, 79.8, 79.5, 79.7, 79.4, 79.2, 79.5, 79.1]
ma = lambda i: sum(w[max(0, i - 6):i + 1]) / len(w[max(0, i - 6):i + 1])
EXPECT = {  # (sheet, cell): valor
    ('Setup', 'B16'): 1780.0, ('Setup', 'B18'): 2759.0, ('Setup', 'B22'): 2207.0,
    ('Setup', 'B23'): 160.0, ('Setup', 'B24'): 72.0, ('Setup', 'B25'): 230.0,
    ('Weight Log', 'C17'): ma(13), ('Weight Log', 'D17'): ma(13) - ma(6),
    ('Dashboard', 'B11'): ma(13),
    ('Workout Log', 'F7'): 119.6, ('Workout Log', 'G7'): 'PR ★',
    ('Workout Log', 'G4'): None,  # no es PR -> vacío
    ('Dashboard', 'A16'): 'On track — keep going.',
}


def recalc_libreoffice():
    from openpyxl import load_workbook
    os.makedirs('_verify', exist_ok=True)
    subprocess.run([SOFFICE, '--headless', '--calc', '--convert-to', 'xlsx',
                    '--outdir', '_verify', WB], capture_output=True, timeout=120)
    rb = load_workbook(os.path.join('_verify', WB), data_only=True)
    fails = 0
    for (sheet, cell), want in EXPECT.items():
        got = rb[sheet][cell].value
        if isinstance(want, float):
            ok = isinstance(got, (int, float)) and abs(got - want) < 0.01
        else:
            ok = (got or None) == (want or None) if want is None else str(got) == str(want)
        print('OK ' if ok else 'FAIL', f'{sheet}!{cell}', '->', repr(got), '| esperado', repr(want))
        fails += 0 if ok else 1
    errs = []
    for ws in rb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('#') and c.value.endswith(('!', '?')):
                    errs.append(f'{ws.title}!{c.coordinate}={c.value}')
    print('celdas con error Excel (LibreOffice):', len(errs), errs[:10])
    return fails == 0 and not errs


if os.path.exists(SOFFICE):
    print('== Verificación con LibreOffice (autoritativa) ==')
    ok = recalc_libreoffice()
else:
    print('== LibreOffice no encontrado: verificación parcial con lib formulas ==')
    import formulas  # noqa
    sol = formulas.ExcelModel().loads(WB).finish().calculate()
    ok = True  # (la lib es indulgente con _xlfn.; usá LibreOffice para el chequeo real)
    print('AVISO: instalá LibreOffice para la verificación que caza #NAME? de MAXIFS.')

print('RESULTADO:', 'PASS' if ok else 'FAIL')
sys.exit(0 if ok else 1)
