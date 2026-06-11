# Genera Adaptive_Fitness_Tracker.xlsx (producto #1, OPORTUNIDADES.md).
# Correr desde esta carpeta: python build_tracker.py
from datetime import date
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

DARK, ACCENT, NOTE, GREEN, BLUE = '1F2937', '2563EB', '6B7280', '008000', '0000FF'
WRAP = Alignment(wrap_text=True, vertical='top')

def F(**kw):
    kw.setdefault('name', 'Arial')
    return Font(**kw)

def title(ws, text, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=text)
    c.font = F(bold=True, size=15, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=DARK)
    c.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[1].height = 30

def section(ws, row, text, col_from=1, col_to=3):
    for col in range(col_from, col_to + 1):
        ws.cell(row=row, column=col).fill = PatternFill('solid', start_color=ACCENT)
    c = ws.cell(row=row, column=col_from, value=text)
    c.font = F(bold=True, size=10, color='FFFFFF')

def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', start_color=DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')

def label(ws, cell, text):
    ws[cell] = text
    ws[cell].font = F(size=10)

def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = F(size=9, italic=True, color=NOTE)

wb = Workbook()
dash = wb.active
dash.title = 'Dashboard'
setup = wb.create_sheet('Setup')
wlog = wb.create_sheet('Weight Log')
xlog = wb.create_sheet('Workout Log')
guide = wb.create_sheet('Guide')
dash.sheet_properties.tabColor = ACCENT
setup.sheet_properties.tabColor = BLUE
guide.sheet_properties.tabColor = '16A34A'

# ---------- Setup ----------
title(setup, 'SETUP — fill in the blue cells', 6)
section(setup, 3, 'YOUR DETAILS')
rows = [('Units', 'Metric', 'Imperial = pounds & inches'),
        ('Sex', 'M', ''),
        ('Age', 30, ''),
        ('Height', 180, 'cm (Metric) or total inches (Imperial)'),
        ('Weight', 80, 'kg (Metric) or lb (Imperial)'),
        ('Activity level', 'Moderate', 'see table on the right'),
        ('Goal', 'Cut', 'Cut -20% | Maintain | Bulk +10%'),
        ('Protein level', 2.0, 'grams of protein per kg of bodyweight')]
for i, (lab, val, hint) in enumerate(rows, start=4):
    label(setup, f'A{i}', lab)
    setup[f'B{i}'] = val
    setup[f'B{i}'].font = F(bold=True, color=BLUE)
    if hint:
        note(setup, f'C{i}', hint)

section(setup, 13, "ENGINE (don't touch)")
engine = [('Weight (kg)', '=IF($B$4="Imperial",$B$8*0.453592,$B$8)', '0.0'),
          ('Height (cm)', '=IF($B$4="Imperial",$B$7*2.54,$B$7)', '0.0'),
          ('BMR', '=10*$B$14+6.25*$B$15-5*$B$6+IF($B$5="M",5,-161)', '0'),
          ('Activity multiplier', '=VLOOKUP($B$9,$E$4:$F$8,2,FALSE)', '0.000'),
          ('TDEE', '=$B$16*$B$17', '0'),
          ('Goal adjustment', '=VLOOKUP($B$10,$E$11:$F$13,2,FALSE)', '0%')]
for i, (lab, formula, fmt) in enumerate(engine, start=14):
    label(setup, f'A{i}', lab)
    setup[f'B{i}'] = formula
    setup[f'B{i}'].number_format = fmt
    setup[f'B{i}'].font = F(size=10)
setup['B16'].comment = Comment('Mifflin-St Jeor equation (1990), the most accurate BMR formula for the general population.', 'Adaptive Fitness Tracker')

section(setup, 21, 'YOUR DAILY TARGETS')
targets = [('Calories (kcal)', '=ROUND($B$18*(1+$B$19),0)'),
           ('Protein (g)', '=ROUND($B$14*$B$11,0)'),
           ('Fat (g)', '=ROUND($B$14*0.9,0)'),
           ('Carbs (g)', '=ROUND(($B$22-$B$23*4-$B$24*9)/4,0)')]
for i, (lab, formula) in enumerate(targets, start=22):
    label(setup, f'A{i}', lab)
    setup[f'B{i}'] = formula
    setup[f'B{i}'].number_format = '0'
    setup[f'B{i}'].font = F(bold=True, size=11)

setup['E3'] = 'Activity'
setup['F3'] = 'Factor'
for c in ('E3', 'F3'):
    setup[c].font = F(bold=True, size=9, color=NOTE)
for i, (act, mult) in enumerate([('Sedentary', 1.2), ('Light', 1.375), ('Moderate', 1.55),
                                 ('Very Active', 1.725), ('Athlete', 1.9)], start=4):
    setup[f'E{i}'] = act
    setup[f'F{i}'] = mult
    setup[f'E{i}'].font = setup[f'F{i}'].font = F(size=9)
setup['E10'] = 'Goal'
setup['F10'] = 'Adj.'
for c in ('E10', 'F10'):
    setup[c].font = F(bold=True, size=9, color=NOTE)
for i, (goal, adj) in enumerate([('Cut', -0.20), ('Maintain', 0.0), ('Bulk', 0.10)], start=11):
    setup[f'E{i}'] = goal
    setup[f'F{i}'] = adj
    setup[f'E{i}'].font = setup[f'F{i}'].font = F(size=9)
    setup[f'F{i}'].number_format = '0%'

for dv_args, cell in [(('list', '"Metric,Imperial"'), 'B4'), (('list', '"M,F"'), 'B5'),
                      (('list', '=$E$4:$E$8'), 'B9'), (('list', '=$E$11:$E$13'), 'B10'),
                      (('list', '"1.6,1.8,2.0,2.2"'), 'B11')]:
    dv = DataValidation(type=dv_args[0], formula1=dv_args[1], allow_blank=False)
    setup.add_data_validation(dv)
    dv.add(setup[cell])

for col, w in (('A', 22), ('B', 13), ('C', 36), ('D', 3), ('E', 14), ('F', 8)):
    setup.column_dimensions[col].width = w

# ---------- Weight Log ----------
title(wlog, 'WEIGHT LOG — same scale, every morning', 5)
header_row(wlog, 3, ['Date', 'Weight', 'Trend (7-day avg)', 'Change / week', '% bodyweight / wk'])
sample_w = [80.0, 80.3, 79.9, 80.1, 79.8, 79.9, 79.6, 79.8, 79.5, 79.7, 79.4, 79.2, 79.5, 79.1]
for r in range(4, 124):
    if r - 4 < len(sample_w):
        wlog[f'A{r}'] = date(2026, 5, 25 + (r - 4)) if r - 4 < 7 else date(2026, 6, r - 10)
        wlog[f'B{r}'] = sample_w[r - 4]
        wlog[f'B{r}'].font = F(color=BLUE)
    window = f'$B$4:$B{r}' if r < 10 else f'$B{r-6}:$B{r}'
    wlog[f'C{r}'] = f'=IF($B{r}="","",IF(COUNT($B$4:$B{r})<3,"",AVERAGE({window})))'
    if r >= 11:
        wlog[f'D{r}'] = f'=IF(OR($C{r}="",$C{r-7}=""),"",$C{r}-$C{r-7})'
        wlog[f'E{r}'] = f'=IF($D{r}="","",$D{r}/$C{r})'
    wlog[f'A{r}'].number_format = 'yyyy-mm-dd'
    for col, fmt in (('B', '0.0'), ('C', '0.00'), ('D', '+0.00;-0.00'), ('E', '+0.0%;-0.0%')):
        wlog[f'{col}{r}'].number_format = fmt
note(wlog, 'G4', 'Sample data in blue — replace with yours (see Guide).')
wlog.freeze_panes = 'A4'
for col, w in (('A', 12), ('B', 10), ('C', 16), ('D', 13), ('E', 16), ('G', 40)):
    wlog.column_dimensions[col].width = w

# ---------- Workout Log ----------
title(xlog, 'WORKOUT LOG — track strength, not soreness', 7)
header_row(xlog, 3, ['Date', 'Exercise', 'Weight', 'Reps', 'RIR', 'e1RM', 'PR?'])
sample_x = [(date(2026, 5, 25), 'Squat', 100, 5, 2), (date(2026, 5, 25), 'Bench Press', 80, 5, 2),
            (date(2026, 5, 28), 'Deadlift', 140, 3, 2), (date(2026, 6, 1), 'Squat', 102.5, 5, 2),
            (date(2026, 6, 1), 'Bench Press', 82.5, 4, 1), (date(2026, 6, 4), 'Deadlift', 145, 3, 2)]
for r in range(4, 204):
    if r - 4 < len(sample_x):
        d, ex, w, reps, rir = sample_x[r - 4]
        for col, v in (('A', d), ('B', ex), ('C', w), ('D', reps), ('E', rir)):
            xlog[f'{col}{r}'] = v
            xlog[f'{col}{r}'].font = F(color=BLUE)
    xlog[f'F{r}'] = f'=IF(OR($C{r}="",$D{r}=""),"",ROUND($C{r}*(1+$D{r}/30),1))'
    # MAXIFS es post-2007: en el XML va con prefijo _xlfn. o Excel/LibreOffice dan #NAME?
    xlog[f'G{r}'] = f'=IF($F{r}="","",IF($F{r}>=_xlfn.MAXIFS($F$4:$F$203,$B$4:$B$203,$B{r}),"PR ★",""))'
    xlog[f'A{r}'].number_format = 'yyyy-mm-dd'
    xlog[f'C{r}'].number_format = '0.0'
    xlog[f'F{r}'].number_format = '0.0'
    xlog[f'G{r}'].font = F(bold=True, color='D97706')
xlog['F3'].comment = Comment('Estimated 1-rep max, Epley formula: weight x (1 + reps/30).', 'Adaptive Fitness Tracker')
xlog['I3'] = 'Exercise list (edit me)'
xlog['I3'].font = F(bold=True, size=9, color=NOTE)
for i, ex in enumerate(['Squat', 'Bench Press', 'Deadlift', 'Overhead Press', 'Barbell Row',
                        'Pull-Up', 'Dip', 'Incline Bench', 'Romanian Deadlift', 'Front Squat',
                        'Hip Thrust', 'Lat Pulldown', 'Seated Cable Row', 'Leg Press',
                        'Bulgarian Split Squat'], start=4):
    xlog[f'I{i}'] = ex
    xlog[f'I{i}'].font = F(size=9)
dv_ex = DataValidation(type='list', formula1='=$I$4:$I$23', allow_blank=True)
xlog.add_data_validation(dv_ex)
dv_ex.add('B4:B203')
xlog.freeze_panes = 'A4'
for col, w in (('A', 12), ('B', 22), ('C', 9), ('D', 7), ('E', 6), ('F', 9), ('G', 11), ('I', 24)):
    xlog.column_dimensions[col].width = w

# ---------- Dashboard ----------
title(dash, 'ADAPTIVE FITNESS TRACKER', 9)
section(dash, 3, 'YOUR DAILY TARGETS')
for i, (lab, ref, unit) in enumerate([('Calories', '=Setup!$B$22', 'kcal'),
                                      ('Protein', '=Setup!$B$23', 'g'),
                                      ('Fat', '=Setup!$B$24', 'g'),
                                      ('Carbs', '=Setup!$B$25', 'g')], start=4):
    label(dash, f'A{i}', lab)
    dash[f'B{i}'] = ref
    dash[f'B{i}'].font = F(bold=True, size=11, color=GREEN)
    dash[f'B{i}'].number_format = '0'
    note(dash, f'C{i}', unit)
label(dash, 'A8', 'Goal')
dash['B8'] = '=Setup!$B$10'
dash['B8'].font = F(bold=True, size=11, color=GREEN)

section(dash, 10, 'WEIGHT TREND')
last = lambda col: (f"=IFERROR(LOOKUP(2,1/('Weight Log'!${col}$4:${col}$123<>\"\"),"
                    f"'Weight Log'!${col}$4:${col}$123),\"—\")")
trend = [('Trend weight (7-day)', last('C'), '0.00'),
         ('Change per week', last('D'), '+0.00;-0.00'),
         ('Rate (% bodyweight/wk)', last('E'), '+0.0%;-0.0%')]
for i, (lab, formula, fmt) in enumerate(trend, start=11):
    label(dash, f'A{i}', lab)
    dash[f'B{i}'] = formula
    dash[f'B{i}'].number_format = fmt
    dash[f'B{i}'].font = F(bold=True, size=11)
dash['B13'].comment = Comment('Weekly rate as % of bodyweight. Healthy cut: -0.25% to -1%. Lean bulk: +0.1% to +0.5%.', 'Adaptive Fitness Tracker')

section(dash, 15, 'COACH SUGGESTION')
dash.merge_cells('A16:C18')
dash['A16'] = ('=IF($B$13="—","Log your weight daily for ~2 weeks to unlock coaching.",'
               'IF(Setup!$B$10="Cut",'
               'IF($B$13>-0.0025,"Trend is flat — drop ~150 kcal (from carbs) or add 2,000 steps/day.",'
               'IF($B$13<-0.01,"Losing too fast — add ~150 kcal back to protect muscle.",'
               '"On track — keep going.")),'
               'IF(Setup!$B$10="Bulk",'
               'IF($B$13<0.001,"Scale is not moving — add ~150 kcal.",'
               'IF($B$13>0.005,"Gaining too fast — trim ~100 kcal.",'
               '"On track — keep going.")),'
               '"Maintenance: stay within ±0.25% per week.")))')
dash['A16'].font = F(size=11, italic=True)
dash['A16'].alignment = WRAP

chart = LineChart()
chart.title = 'Weight vs 7-day trend'
chart.style = 12
chart.displayBlanksAs = 'gap'
chart.height, chart.width = 9, 17
data = Reference(wlog, min_col=2, min_row=3, max_col=3, max_row=123)
chart.add_data(data, titles_from_data=True)
chart.set_categories(Reference(wlog, min_col=1, min_row=4, max_row=123))
dash.add_chart(chart, 'E3')
for col, w in (('A', 24), ('B', 12), ('C', 8)):
    dash.column_dimensions[col].width = w

# ---------- Guide ----------
title(guide, 'GUIDE — read me first', 1)
guide.column_dimensions['A'].width = 110
content = [
    (3, 'GET STARTED IN 3 STEPS', True),
    (4, '1. Open the Setup tab and fill in the blue cells (units, sex, age, height, weight, activity, goal).', False),
    (5, '2. Weigh yourself every morning (after bathroom, before food or drink) and log it in Weight Log.', False),
    (6, '3. Log your main lifts in Workout Log. Check the Dashboard once a week — not every day.', False),
    (8, 'WHY TREND WEIGHT (AND NOT THE SCALE)', True),
    (9, 'Daily weight swings 1-2% from water, sodium, carbs and stress. The 7-day moving average filters that noise and shows your real direction. Judge progress only by the trend line.', False),
    (11, 'WHAT IS RIR?', True),
    (12, 'Reps In Reserve: how many clean reps you had left in the tank. RIR 0 = absolute failure, RIR 2 = could have done 2 more. Most working sets should live at RIR 1-3.', False),
    (13, 'The tracker turns every set into an estimated 1-rep max (Epley formula). If your e1RM creeps up over the weeks, you are progressing — even when the weight on the bar looks the same.', False),
    (15, 'WHEN TO ADJUST CALORIES', True),
    (16, 'Never react to a single weigh-in. Collect at least 2 full weeks of data, then follow the Coach suggestion on the Dashboard. Adjust in ~150 kcal steps, taken from carbs, and hold for another 2 weeks.', False),
    (18, 'SAMPLE DATA', True),
    (19, 'Weight Log and Workout Log ship with 2 weeks of sample data (blue cells) so you can see how everything works. Delete only the blue entries — never the formula columns — and start logging.', False),
    (21, 'COMPATIBILITY', True),
    (22, 'Works in Microsoft Excel 2019+/365 and Google Sheets (upload to Drive, then Open with Google Sheets). The PR column uses MAXIFS, which needs Excel 2019 or newer.', False),
    (24, 'DISCLAIMER', True),
    (25, 'This template is for educational purposes and is not medical or dietary advice. Consult a professional before major changes, especially with existing health conditions.', False),
]
for row, text, is_header in content:
    if is_header:
        section(guide, row, text, 1, 1)
    else:
        guide[f'A{row}'] = text
        guide[f'A{row}'].font = F(size=10)
        guide[f'A{row}'].alignment = WRAP
        guide.row_dimensions[row].height = 28

# ---------- Print setup (que el comprador imprima/exporte limpio) ----------
def fit_wide(ws, tall=1):
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = tall
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5

from openpyxl.worksheet.properties import PageSetupProperties
for ws in (dash, setup, wlog, xlog, guide):
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

dash.page_setup.orientation = 'landscape'
fit_wide(dash, tall=1)
dash.print_area = 'A1:N20'
fit_wide(setup, tall=1)
setup.print_area = 'A1:F26'
fit_wide(wlog, tall=0)
wlog.print_title_rows = '3:3'
fit_wide(xlog, tall=0)
xlog.print_title_rows = '3:3'
fit_wide(guide, tall=1)
guide.print_area = 'A1:A25'

wb.save('Adaptive_Fitness_Tracker.xlsx')
print('OK: Adaptive_Fitness_Tracker.xlsx')
